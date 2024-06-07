import fetch_uniprot
import re, warnings, time, openpyxl, traceback

from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from Bio.PDB import PDBParser
from copy import deepcopy
from QOL_functions import *
import pandas as pd
import numpy as np
from sklearn.neighbors import BallTree
from operator import itemgetter
#classes OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
class AnAtom:
    def __init__(self, element='', name='', location=None, residue = None, b_factor = None, consfScore = None, colorScore = None, number = None):
        self.element        = element
        self.name           = name
        self.location       = location
        self.residue        = residue
        self.b_factor       = b_factor
        self.consfScore     = consfScore
        self.colorScore     = colorScore
        self.number         = number
        self.burial             = None          #determined in calculate_burials
        self.normBurial         = None          #determined in calculate_burials
        self.isCavity           = None          #determined in crawl_cavities
        self.normBurials        = {}
        self.cavity             = None          #determined in ACavity initialization in crawl_cavities
    def __repr__(self):
        return f"atom {self.name} in {self.residue}"
    def transform(self, matrix= None):
        """ Apply a 3x4 transformation matrix to an atom's location. return new atom. """
        x, y, z = self.location
        new_x = matrix[0][0] * x + matrix[0][1] * y + matrix[0][2] * z + matrix[0][3]
        new_y = matrix[1][0] * x + matrix[1][1] * y + matrix[1][2] * z + matrix[1][3]
        new_z = matrix[2][0] * x + matrix[2][1] * y + matrix[2][2] * z + matrix[2][3]
        transformed_location = np.array([new_x, new_y, new_z])
        transformed_atom = AnAtom(element=self.element, name=self.name,
                                  location=transformed_location, b_factor = self.b_factor,
                                  consfScore = self.consfScore, colorScore = self.colorScore, number = self.number)

        return transformed_atom
class AResidue:
    def __init__(self, name='', atoms = None, number = None, consfScore = None, colorScore = None):
        self.name       = name
        self.atoms      = atoms if atoms else []
        self.number     = number
        self.consfScore = consfScore
        self.colorScore = colorScore

        self.burial             = None              #determined in calculate_burials
        self.normalizedBurial   = None              #determined in calculate_burials
        self.isCavity           = None              #determined in crawl_cavities

        self.chain      = None          #assigned in AChain.adopt()
        self.location   = None          #assigned after all atoms are adopted in make_asym_units() and AnAsymUnit.make_transformants()

        self.sites      = {}            #determined in find_activeSites()

    def __repr__(self):
        return f"res {self.name}{self.number}"
    def adopt(self, atom: AnAtom):
        """ Add an atom to the residue. """
        self.atoms.append(atom)
        atom.residue = self
class AChain:
    def __init__(self,  name='', indexMap = None, consfMap = None,
                 percMismatch = None, percGap = None, residues=None, goodChain = False, targetChain = None):

        self.name           = name
        self.residues       = residues if residues else []
        self.goodChain      = goodChain                             #only goodChains are analyzed, determined/assigned in make_asym_units_get_index_map()
                                                                        # and assigned again in AnAsymUnit.make_transformants()
        self.targetChain    = targetChain                           #chain is not DNA or belongs to different gene, determined/assigned in make_asym_units_get_index_map()
        self.asymUnit       = None                              #determined/assigned as above, but only after the chain is already made

        self.indexMap       = indexMap
        self.consfMap       = consfMap
        self.percMismatch   = percMismatch
        self.percGap        = percGap

    def __repr__(self):
        return f"{self.name}"
    def adopt(self, residue: AResidue):
        """ Add a residue to the chain. """
        self.residues.append(residue)
        residue.chain = self
class AnAsymUnit:
    def __init__(self, chains=None):
        self.chains             = chains if chains else []
        self.solvents           = []
        self.other_heteroatoms  = []
    def make_transformants(self, transformation_matrices, asym_units):
        """ Apply transformations on self to generate additional asymmetric units with renamed chains.
        asym_units is only passed to determine available chain names"""
        taken_chain_labels      = gaze_range(AChain, asym_units, 'name')
        chain_labels            = iter([i for i in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ' if i not in taken_chain_labels])
        child_Asyms = []
        for matrix in transformation_matrices:              # for transformation matrix
            if matrix == [[1.0, 0.0, 0.0, 0.0], [0.0, 1.0, 0.0, 0.0], [0.0, 0.0, 1.0, 0.0]]:
                child_Asyms.append(self)
            else:
                child_Asym = AnAsymUnit()                           # make a child_Asym
                for parent_chain in self.chains:                        # for parent_chain in parent_Asym
                    child_chain_id = next(chain_labels, 'Z')                # Assign new child_chain label, default to 'Z' if run out of labels
                    child_chain = AChain(name=child_chain_id, goodChain = parent_chain.goodChain)               # make new child_chain
                    for parent_residue in parent_chain.residues:            # for parent_residue in parent_chain
                        child_residue = AResidue(name=parent_residue.name,
                                                 number=parent_residue.number, consfScore=parent_residue.consfScore, colorScore=parent_residue.colorScore)
                        for parent_atom in parent_residue.atoms:                # for parent_atom in parent_residue
                            child_atom          = parent_atom.transform(matrix)
                            child_atom.residue  = child_residue
                            child_residue.adopt(child_atom)                 # give child_atom to child_residue
                        child_residue.location = average_location(child_residue.atoms)
                        child_chain.adopt(child_residue)              # give child_residue to child_chain
                    child_Asym.adopt(child_chain)                   # give child_chain to child_Asym
                for parent_atom in self.solvents:                            # for solvent in parent_Asym
                    child_atom = parent_atom.transform(matrix)              # transform atom
                    child_Asym.solvents.append(child_atom)                  # give atom to child_residue
                for parent_atom in self.other_heteroatoms:                   # for other_heteroatom in parent_Asym
                    child_atom = parent_atom.transform(matrix)              # transform atom
                    child_Asym.other_heteroatoms.append(child_atom)         # give atom to child_asym
                child_Asyms.append(child_Asym)                          # save child_Asym in child_Asyms

        return child_Asyms                                                  # return child_Asyms
    def adopt(self, chain: AChain):
        """ Add a chain to the asymmetric unit. """
        self.chains.append(chain)
        chain.AsymUnit = self
class AnActiveSite:
    def __init__(self, residues, location):
        self.residues   = residues
        self.location   = location
        self.consfScore = sum([residue.consfScore for residue in self.residues])
    def __repr__(self):
        residues = [residue.name+str(residue.number) for residue in self.residues]
        chain    = set([residue.chain for residue in self.residues])
        return f"activeSite {residues} in chain {chain} at {self.location}"
    def __str__(self):
        return self.__repr__()
class ACavity:
    def __init__(self, pAtoms, atomLocTree):
        self.volume         = len(pAtoms)
        self.pAtoms         = pAtoms
        self.pAtomLocTree   = make_LocTree(list(pAtoms))
        nearbyAtomBins      = [find_close_objects(pAtom, atomLocTree,3) for pAtom in pAtoms]
        self.atoms          = set(atom for atomBin in nearbyAtomBins for atom in atomBin)
        self.residues       = set(atom.residue for atom in self.atoms)
        self.conservedSites = set(residue.sites.get(8, None) for residue in self.residues)
        self.activeSites    = set(residue.sites.get(9, None) for residue in self.residues)
        self.conservedSites .discard(None)
        self.activeSites    .discard(None)

        self.activeSiteScore    = sum([activeSite.consfScore for activeSite in self.activeSites])
        self.conservedSiteScore = sum([conservedSite.consfScore for conservedSite in self.conservedSites])
        self.consfScore         = sum(sigNorm([residue.consfScore for residue in self.residues]))

        for pAtom in pAtoms:
            pAtom.cavity = self
        for atom in self.atoms:
            atom.cavity = self
            atom.isCavity = True

# broad use functions OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
def fetch(object_type, objects, attribute_name = None, attribute_value = None, relation=None):
    """
    Recursively fetch objects of a specified type. If property is provided, returns objects for which property is true. else, return objects of object type.

    :param object_type: The type of objects to retrieve (AnAsymUnit, AChain, AResidue, AnAtom).
    :param objects: A list of objects to search through.
    :param property: string of a bool in the form 'obj.attribute_name == value' could also be <=, in, etc. must use 'obj' regardless of object_type
    :return: A list of matched objects.
    """

    def relate(obj1, obj2, relation):
        """
        Check if a specified relation holds between two objects using a dictionary of lambda functions.

        :param obj1: First object.
        :param obj2: Second object.
        :param relation: String representing the relation ('==', '<', '>', 'in', '<=', '>=', '!=').
        :return: Boolean indicating whether the relation holds.
        """
        operations =    {'<': lambda x, y: x < y,
                        '>': lambda x, y: x > y,    'in': lambda x, y: x in y,
                        '<=': lambda x, y: x <= y,  '>=': lambda x, y: x >= y,
                        '!=': lambda x, y: x != y, '==': lambda x, y: x == y}

        if relation is None:
            return obj1 == obj2
        if relation not in operations:
            raise ValueError(f"Unsupported relation: {relation}")
        return operations[relation](obj1, obj2)

    attr_to_nested  = { AnAsymUnit: 'chains', AChain: 'residues',
                        AResidue: 'atoms', AnActiveSite: 'residues', AnAtom: []}                #name of attribute referring to list of nested object for each type
    matched_objects = []                                                                         #matches (to return)

    for obj in objects:                                                                          #for object in objects
        if isinstance(obj, object_type):                                                            #if object is of type object_type
            if attribute_name and relate(getattr(obj,attribute_name),attribute_value,relation):         #if property holds for object
                matched_objects.append(obj)                                                                 #add object to matched_objects
            elif not attribute_name:                                                                          #elif no property is given ("requested only that type is object_type")
                matched_objects.append(obj)                                                                 #add object to matched_objects

        else:                                                                                       #else object is not of type object_type
            nested_objects                  = getattr(obj, attr_to_nested[type(obj)], obj)                                   #get its nested objects, e.g. chain:residues, residue:atoms
            generational_nested_matches     = fetch(object_type,nested_objects, attribute_name, attribute_value, relation)  #subcall this function on the nested objects
            matched_objects                 .extend(generational_nested_matches)                                            #subcall returns matches, store them with this call's matches
    return matched_objects                                                                  #return matches
def gaze_range(object_type, objects, attribute_name):
    """
    return a list of values for a attribute_name from all objects of type object_type in objects and their sub_objects

    :param object_type: The type of objects to check ('AnAsymUnit', 'AChain', 'AResidue', 'AnAtom').
    :param objects: A list of objects to search through.
    :param attribute_name: The name of the attribute to extract values from.
    :return: A list of attribute values.
    """
    fetched_objects     = fetch(object_type, objects)
    attribute_values    = [getattr(obj, attribute_name, f'failed to fetch {attribute_name} of {obj}') for obj in fetched_objects]
    return attribute_values
def get_distance(obj1, obj2):
    """
    Calculate the Euclidean get_distance between two objects.

    :param obj1: First atom with 'location' attribute as a list [x, y, z].
    :param obj2: Second atom with 'location' attribute as a list [x, y, z].
    :return: Distance between obj1 and obj2.
    """

    location1 = np.array(obj1) if isinstance(obj1, list) else obj1.location
    location2 = np.array(obj2) if isinstance(obj2, list) else obj2.location
    return np.linalg.norm(location1 - location2)
def make_LocTree(objects):
    """
    Generate a ball tree from a list of locations.

    :param locations: List of locations, where each location is a tuple or list of coordinates.
    :return: Ball tree object.
    """
    if type(objects[0]) == tuple:
        locations = [tuple(round(float(dimension),4) for dimension in object)           for object in objects]
    else:
        locations = [tuple(round(float(dimension),4) for dimension in object.location)  for object in objects]
    locationDict = {location:objects[i] for i, location in enumerate(locations)}
    # Convert the locations to a numpy array
    data = np.array(locations)

    # Create a ball tree using the data
    ball_tree = BallTree(data)

    return [ball_tree, locationDict]
def find_close_objects(target, treePack, range_cutoff, isArray = False):
    """
    Find all objects within a specified range of a target location using a ball tree.

    :param ball_tree: Ball tree object.
    :param target_location: Target location to search around.
    :param range_cutoff: Maximum distance from the target location to consider.
    :return: List of locations within the specified range.
    """
    ball_tree, locationDict = treePack
    if not isArray:
        target_location = target.location
    else:
        target_location = target
    # Convert the target location to a numpy array
    target_data = np.array([target_location])
    # Perform the range search using the ball tree
    indices = ball_tree.query_radius(target_data, r=range_cutoff)[0]
    # Iterate over the indices and retrieve the corresponding locations
    locations = []
    for index in indices:
        location = ball_tree.data[index]
        locations.append(tuple(round(dimension, 4) for dimension in location))
    objects = set(locationDict[location] for location in locations)
    return objects
def average_location(objects):
    """ Calculate the average location of a list of atoms or residues. """
    x = sum(obj.location[0] for obj in objects) / len(objects)
    y = sum(obj.location[1] for obj in objects) / len(objects)
    z = sum(obj.location[2] for obj in objects) / len(objects)

    return np.array([x,y,z])
def sigNorm(consfScores):
    """normalizes a list of consfScores using the sigmoid function to
        minimize the contribution from unconserved residues"""
    arConsfScores = np.array(consfScores)+2
    sigNorm = 1 / (1 + np.exp(arConsfScores))  # sigmoid minimizes weight of low scores (now low conservations)
    normScores = sigNorm * consfScores  # scores normalized by their sigmoid value
    return list(normScores)
def normalize_aList(aList):
    """normalize a list of values to a 0-1 scale"""
    arList = np.array(aList)
    normList = (arList - arList.min()) / (arList.max() - arList.min())
    return list(normList)

# body functions OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
def iterate_structures(all_pdbIDs, projDir, pdbDir, consfDicts, targetSeq, uniprotID, consfID,
                       cavitySearch, defaultStructData, targetChainName, resetCavities, getColorCmds):
    def process_structure(pdbID, projDir, pdbDir, consfDicts, structData, defaultStructData, targetChainName, resetCavities):
        def make_asym_units_get_index_map(pdbFileName, projDir, pdbDir, targetSeq, pdbID, consfDicts, structData):
            """ Read a PDB file using PDBParser and extract the structure into nested classes defined above, in a list asymUnits.
                generate indexMap, dict mapping indexes in structure to indexes in targetSeq.
                update structData with structureName, %mismatch, and %gap
                return:
                asymUnits, indexMap, structData"""

            if not is_in_Dir(pdbFileName, pdbDir):
                issue(f'cannot process {pdbID}, file does not exist.', projDir)
                return None, structData

            def find_validChains_chainVars(chains, targetSeq, pdbID, projDir):
                validChains = []
                chainVars   = {}
                validChaini = 0
                ThreeToOneLetter = {'ALA': 'A', 'ARG': 'R', 'ASN': 'N', 'ASP': 'D',
                                    'CYS': 'C', 'GLN': 'Q', 'GLU': 'E', 'GLY': 'G',
                                    'HIS': 'H', 'ILE': 'I', 'LEU': 'L', 'LYS': 'K',
                                    'MET': 'M', 'PHE': 'F','PRO': 'P', 'SER': 'S',
                                    'THR': 'T', 'TRP': 'W', 'TYR': 'Y', 'VAL': 'V'}

                for chaini, chain in enumerate(chains):

                    #check if a chain is valid
                    residues,indexes = '',[]
                    problemChain     = False
                    for resi, residue in enumerate(chain):
                        res_category = residue.id[0]
                        if res_category == " ":
                            res_name        = residue.get_resname()
                            res_number      = residue.id[1]
                            res_one_letter  = ThreeToOneLetter.get(res_name, ' ')
                            try:
                                residues        += res_one_letter               #in problemchains, res_one_letter is None even though it is impossible to produce such a value
                                indexes         .append(res_number)
                            except:
                                problemChain = True
                                issue(f'issue assigning a residue to {res_name}{resi}', projDir)

                    #exclude chains with unmapped residues
                    if problemChain:
                        issue(f'Issue processing chain "{chain.id}" in {pdbID}: invalid residue {res_name} at {res_number}', projDir)
                        continue

                    #if not DNA
                    if residues.strip() != '':
                        indexMap, consfMap, percMismatch, percGap   = fetch_uniprot. get_index_consf_maps_percMismatch(
                            oriIdealSeq=targetSeq, oriStructSeq=residues, structIndexes=indexes)

                        #exclude mismatching chains (heterocomplexes)
                        if percMismatch < 80 and percGap < 70:
                            chainVars[validChaini]   = {'indexMap'    :indexMap,
                                                        'consfMap'     :consfMap,
                                                        'percMismatch' :percMismatch,
                                                        'percGap'      :percGap,
                                                        'targeted'     :True}
                            validChains .append(chain)
                            validChaini += 1
                        else:
                            issue(f'chain "{chain.id}" in {pdbID}: chain does not match uniprot target '
                                  f'mismatch: {percMismatch}% '
                                  f'gap: {percGap}%'
                                  '\nchain will not receive conservation analysis', projDir)
                            chainVars[validChaini] = {'indexMap': indexMap,
                                                      'consfMap': consfMap,
                                                      'percMismatch': percMismatch,
                                                      'percGap': percGap,
                                                      'targeted': False}
                            validChains.append(chain)
                            validChaini += 1
                            continue

                    #exclude DNA-only chains
                    else:
                        issue(f'Issue processing chain "{chain.id}" in {pdbID}: no residues found'
                              f'\nchain will not receive conservation analysis', projDir)
                        chainVars[validChaini] = {'indexMap': {},
                                                      'consfMap': {},
                                                      'percMismatch': '',
                                                      'percGap': '',
                                                      'targeted': False}
                        validChains.append(chain)
                        validChaini += 1
                        continue

                #if no valid chains at all
                if not any([chainVar['targeted'] for chainVar in chainVars.values()]):
                    issue(f'OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO'
                          f'No valid chains found in {pdbID}, treating all chains as valid! This may invalidate analysis!!\n'
                          'OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO', projDir)
                    for chainVar in chainVars.values():
                        chainVar['targeted'] = True

                return validChains, chainVars
            def make_AnAsymUnit(validChains, chainVars, consfDicts):
                def reindex_pre_Met0s(residues):
                    """reindexes unmapped residues in N-root"""
                    for i, residue in enumerate(residues):
                        if residue.number is not None:
                            firsNumber = residue.number
                            for j in range(i, -1, -1):
                                residues[j].number = firsNumber
                                firsNumber -= 1
                            break

                ThreeToOneLetter = {'ALA': 'A', 'ARG': 'R', 'ASN': 'N', 'ASP': 'D',
                                    'CYS': 'C', 'GLN': 'Q', 'GLU': 'E', 'GLY': 'G',
                                    'HIS': 'H', 'ILE': 'I', 'LEU': 'L', 'LYS': 'K',
                                    'MET': 'M', 'PHE': 'F', 'PRO': 'P', 'SER': 'S',
                                    'THR': 'T', 'TRP': 'W', 'TYR': 'Y', 'VAL': 'V'}

                consfScores, colorScores, consfPositions, _ = consfDicts if consfDicts else \
                    [{None: None} for _ in range(3)] + [{'%mismatch': None, '%gap': None}]

                thisAsymUnit    = AnAsymUnit()                                        # make asym_unit
                chainResolBins   = {}
                for chaini, chain in enumerate(validChains):                        # for chain in model
                    targetChain     = chainVars[chaini]['targeted']
                    indexMap        = chainVars[chaini]['indexMap']
                    consfMap        = chainVars[chaini]['consfMap']
                    percMismatch    = chainVars[chaini]['percMismatch']
                    percGap         = chainVars[chaini]['percGap']

                    thisAChain      = AChain(chain.id, indexMap, consfMap, percMismatch, percGap,
                                             targetChain = targetChain)                                                 # make AChain
                    for resi, residue in enumerate(chain):                                                              # for residue in chain
                        res_category    = residue.id[0]                                                                     # get residue category (standard or HETATM)
                        res_name        = residue.get_resname()                                                             # get residue name
                        res_struct_num  = residue.id[1]                                                                     # get number in pdb file
                        if res_category == " ":                                                                             # if standard residue
                            res_one_letter  = ThreeToOneLetter  .get(res_name, None)                                            # get one letter code
                            resNumber       = indexMap          .get(res_struct_num, None)                                      # recalculate from pdb number to aligned number
                            consfNumber     = consfMap          .get(res_struct_num, None)                                      # recalculate from pdb number to consf number
                            consfResidue    = consfPositions    .get(consfNumber, None)                                         # get residue at same pos in consf
                            if consfResidue != None and consfResidue == res_one_letter:                                         # if residue isn't mutant
                                consfScore = consfScores.get(resNumber,999)                                                         # assign consfScore
                                colorScore = colorScores.get(resNumber,999)                                                         # assign colorScore
                            else:                                                                                               # else if residue is mutant
                                consfScore = 999                                                                                    # assign default consfScore
                                colorScore = 999                                                                                    # assign default colorScore
                            thisAResidue = AResidue(name=res_name, number=resNumber,
                                                   consfScore=consfScore, colorScore=colorScore)                                # make AResidue and assign above variables
                            for atom in residue:                                                                                # for atom in residue
                                thisAnAtom = AnAtom(element=atom.element, name=atom.get_name(),                                       # make atom and assign above variables
                                                  location=np.array(atom.get_coord()), residue = thisAResidue,
                                                  b_factor = atom.get_bfactor(), consfScore=consfScore,
                                                  colorScore = colorScore, number = atom.get_serial_number())
                                thisAResidue.adopt(thisAnAtom)                                                                         # give atom to residue                                                                        #for atom in residue
                            thisAResidue.location = average_location(thisAResidue.atoms)                                          # calc location of residue from atoms
                            thisAChain            .adopt(thisAResidue)                                                            # give residue to chain
                        else:                                                                                               # else if HETATM, not a residue
                            for atom in residue:                                                                                # for atom in molecule
                                thisAnAtom = AnAtom(element=atom.element,                                                         # make AnAtom
                                                  name=atom.get_name(), location=np.array(atom.get_coord()))
                                if res_name == "HOH":  # Water molecule                                                         # if water molecule
                                    thisAsymUnit.solvents.append(thisAnAtom)                                                          # give atom to solvents
                                else:                                                                                           # else (metal ions, molecules, etc)
                                    thisAsymUnit.other_heteroatoms.append(thisAnAtom)                                                 # give atom to other_heteroatoms
                    reindex_pre_Met0s(thisAChain.residues)                                                                  #renumber pre-MET0 residues from None to negatives
                    thisAsymUnit.adopt(thisAChain)                                         # give chain to asym_unit
                return thisAsymUnit
            def assign_goodChain(AnAsymUnit):
                """determine which chains have best terminal resolution"""
                #bin chains by terminal resolution
                resolBins = {}
                for chain in AnAsymUnit.chains:
                    if chain.targetChain:
                        NUnres               = abs(chain.residues[0].number - 1)
                        CUnres               = chain.indexMap[9E100]
                        chainQual            = sum((NUnres, CUnres))
                        resolBins[chainQual] = resolBins[chainQual] if chainQual in resolBins else []   #initiate bin if new
                        resolBins[chainQual] .append(chain)                                             #add chain to bin

                #find best resolution chains
                bestResolChains = resolBins[min(resolBins)] if resolBins else AnAsymUnit.chains
                if not bestResolChains:
                    issue('OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO\n'
                          f'failed to find any chains with resolved termini in {pdbID}, defaulting to all chains\n'
                          'This is likely a domain-specific structure, check structure name in termini_scores.xlsx\n'
                          'This will also be apparent in residue identities in this file.\n'
                          'if you do not intend to do a domain-specific analysis, exclude this result from your analysis\n'
                          'OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO', projDir)
                    bestResolChains = AnAsymUnit.chains
                for chain in bestResolChains:
                    chain.goodChain = True
                #could consider %mism and %gap but its difficult to predict
                # how that would affect the chain itself vs adjacent chains

            parser                      = PDBParser(PERMISSIVE=1)
            pdbData                     = parser.get_structure('PDB_structure', pdbFileName)
            header                      = pdbData.header
            structureName               = header['name'] if 'name' in header else '(no title found)'
            structData['structureName'] = structureName

            #turn valid chains in PDB into AChains
            AnAsymUnits = []
            for asymmetricalUnit in pdbData:
                chains                      = [chain for chain in asymmetricalUnit if chain.id != '_']                  #exclude swissmodel new chains
                validChains, chainVars      = find_validChains_chainVars(chains, targetSeq, pdbID, projDir)
                thisAnAsymUnit              = make_AnAsymUnit(validChains, chainVars, consfDicts)
                AnAsymUnits                 .append(thisAnAsymUnit)
                assign_goodChain(thisAnAsymUnit)

            return AnAsymUnits, structData
        def extract_matrices_oligo_gmqe_qsqe(projDir, pdbFileName, structData):
            """ Extract transformMatrices and oligomeric state (to determine if matrices need to be applied) from a PDB file.
            searches first for biological oligomer, then for software determined oligomer, prints notice if software determined is used.
            return:
            transformMatrices, biologicalUnit"""
            
            pyName          = os.path.basename(__file__)
            pdb_data        = open(pdbFileName, 'r').read()
            biomtPattern   = re.compile(r"^REMARK 350 {3}BIOMT(\d) +(\d) +([ \d.\-]+) +([ \d.\-]+) +([ \d.\-]+) +([ \d.\-]+)")
            unitPattern     = re.compile(r"^REMARK 350 AUTHOR DETERMINED BIOLOGICAL UNIT: (.+)")
            backupPattern   = re.compile(r"^REMARK 350 SOFTWARE DETERMINED QUATERNARY STRUCTURE: (.+)")
            gmqePattern     = re.compile(r"gmqe: (.+)")
            qsqePattern     = re.compile(r"qsqe: (.+)")
            #find oligomer
            biologicalUnit  = ''
            softwareUnit    = ''
            gmqe            = ''
            qsqe            = ''
            for line in pdb_data.splitlines():
                # Check for biological unit
                unitMatch   = unitPattern   .match(line)
                backupMatch = backupPattern .match(line)
                gmqeMatch   = gmqePattern   .match(line)
                qsqeMatch   = qsqePattern   .match(line)
                if unitMatch:
                    biologicalUnit = unitMatch.group(1).strip()
                    break
                if backupMatch:
                    softwareUnit = backupMatch.group(1).strip()
                if gmqeMatch:
                    gmqe = gmqeMatch.group(1).strip()
                if qsqeMatch:
                    qsqe = qsqeMatch.group(1).strip()
            if not biologicalUnit:
                if not softwareUnit:
                    issue(f"PDB did not state either biological or software-determined quaternary structure.\n"
                          f"Defaulting to generate asymmetric unit using transformation matrices.", projDir)
                    biologicalUnit = 'ICOSAMERIC'                                       #default to insufficient chain number
                else:
                    biologicalUnit = softwareUnit

            #find matrices
            transformMatrices = {}
            for line in pdb_data.splitlines():
                match = biomtPattern.match(line)
                if match:
                    row_number, matrix_number, *matrix_values = match.groups()
                    matrix_number = int(matrix_number)
                    row_number = int(row_number) - 1  # Convert to 0-based index
                    matrix_values = [float(val) for val in matrix_values[0].split()]

                    if matrix_number not in transformMatrices:
                        transformMatrices[matrix_number] = [[0, 0, 0, 0] for _ in range(3)]
                    transformMatrices[matrix_number][row_number] = matrix_values
            transformMatrices = list(transformMatrices.values())
            structData['swissmodel GMQE'] = gmqe
            structData['swissmodel QSQE'] = qsqe
            return transformMatrices, biologicalUnit, structData
        def transform_asyms(origAsyms, transformMatrices, oligo):
            """checks if chain number matches oligomer, if not, applies transformation matrices to generate new asymmetric units"""
            if transformMatrices:
                oligomer_to_number  = {"MONOMERIC": 1, "DIMERIC": 2, "TRIMERIC": 3, "TETRAMERIC": 4, "PENTAMERIC": 5,        #standard
                                      "HEXAMERIC": 6,
                                      "HEPTAMERIC": 7, "OCTAMERIC": 8, "NONAMERIC": 9, "DECAMERIC": 10, "UNDECAMERIC": 11,
                                      "DODECAMERIC": 12, "TRIDECAMERIC": 13, "TETRADECAMERIC": 14, "PENTADECAMERIC": 15,
                                      "HEXADECAMERIC": 16, "HEPTADECAMERIC": 17, "OCTADECAMERIC": 18, "NONADECAMERIC": 19,
                                      "ICOSAMERIC": 20
                                       }  # to check if structure needs transforming
                chainNumber         = len([chain for chain in fetch(AChain, origAsyms) if chain.name not in '_-'])             #because swissmodel re-identifies heteroatoms to a different chain
                oligoToNumber       = oligomer_to_number.get(oligo,None)
                oligoToNumber       = oligoToNumber if oligoToNumber != None else re.findall(r'\d+', oligo)[0]
                if chainNumber != oligoToNumber:  # if number of chains matches oligomer
                    transformed_asym_unit_bins = []
                    for asym in origAsyms:
                        transformants = asym.make_transformants(transformMatrices, asym_units=origAsyms)
                        transformed_asym_unit_bins.append(transformants)
                    transformed_asym_unit_bins = [item for sublist in transformed_asym_unit_bins for item in sublist]
                    return transformed_asym_unit_bins
                else:
                    return origAsyms
            else:
                issue('failed to find a transformation matrix, defaulting to written asymmetrical unit.', projDir)
                return origAsyms
        def find_targetChain(AnAsymUnits, targetChainName, structData):
            """find the chain with the most residues"""
            targetChains = fetch(AChain, AnAsymUnits, 'goodChain', True)
            if targetChainName:
                targetChain = fetch(AChain, targetChains, 'name', targetChainName)[0]
            else:
                targetChain = targetChains[0]
            printWrite(f'analyzing chain {targetChain.name} of {len(targetChains)} optimal chains.', projDir)
            structData['chain ID']= targetChain.name
            return targetChain, structData
        def extract_unresolved_termini(targetChain, structData):
            """ Find how many residues are unresolved at the termini of the longest chain. """
            result = {'N unresolved residues':targetChain.residues[0].number - 1,
                        'C unresolved residues':targetChain.indexMap[9E100],}
            structData.update(result)
            return structData
        def _calc_burials(targetAtoms, atomLocTree, cutoff, excludeOwn = False):
            """ Calculate burials for atoms, to be normalized downstream
            used in both calculate_burials() and crawl_cavities()"""
            burials = []
            for i, targetAtom in enumerate(targetAtoms):                                                                #for each CA atom in the chain
                allCloseAtoms            = find_close_objects(targetAtom, atomLocTree, cutoff)                        #find atoms within cutoff dist
                if excludeOwn:
                    ownAtoms        = set(targetAtom.residue.atoms)                                             #remove self from close atoms
                    allCloseAtoms   = allCloseAtoms.difference(ownAtoms)
                burial              = len(allCloseAtoms)                                                            #sum scores
                burials             .append(burial)                                                         #add to burial scores
            return burials
        def calculate_burials(targetChain, burialDistanceThreshold, AnAsymUnits, structData, atomLocTree):
            """ Calculate burial scores for CA atoms in a chain"""
            CAatoms           = fetch(AnAtom, targetChain.residues, 'name', 'CA')                                       #fetch all CA atoms in residues
            allAtoms          = fetch(AnAtom, AnAsymUnits)                                                               #backbone atoms are checked for burial against all atoms
            residues          = fetch(AResidue, targetChain.residues)
            threeToOneLetter  = {None: None, 'ALA': 'A', 'ARG': 'R', 'ASN': 'N', 'ASP': 'D', 'CYS': 'C', 'GLN': 'Q', 'GLU': 'E', 'GLY': 'G',
                               'HIS': 'H', 'ILE': 'I', 'LEU': 'L', 'LYS': 'K', 'MET': 'M', 'PHE': 'F', 'PRO': 'P', 'SER': 'S', 'THR': 'T',
                               'TRP': 'W', 'TYR': 'Y', 'VAL': 'V'}

            def assign_normBurials(allAtoms, cutoff, atomLocTree):
                burials    = _calc_burials(allAtoms, atomLocTree, 5)
                normBurls       = normalize_aList(burials)

                for i, atom in enumerate(allAtoms):
                    atom.burial             = burials[i]
                    atom.normBurial         = normBurls[i]
            assign_normBurials(allAtoms, burialDistanceThreshold, atomLocTree)                                     #calculate burial scores

            def assign_residue_burials(AnAsymUnits):
                residues = fetch(AResidue, AnAsymUnits)
                for residue in residues:
                    residue.burial   = sum([atom.burial for atom in residue.atoms]) / len(residue.atoms)
            assign_residue_burials(AnAsymUnits)

            def assign_residue_normBurials(AnAsymUnits):
                residues = fetch(AResidue, AnAsymUnits)
                normalizedBurials         = normalize_aList([residue.burial for residue in residues])
                for resi, residue in enumerate(residues):
                    residue.normBurial   = normalizedBurials[resi]
            assign_residue_normBurials(AnAsymUnits)

            resNormBurials = [residue.normBurial for residue in targetChain.residues]
            resBurials     = [round(residue.burial, 2) for residue in targetChain.residues]

            termBurials     = [round(residues[0].burial, 2), round(residues[-1].burial, 2)]
            termNormBurials = [round(residues[0].normBurial, 2), round(residues[-1].normBurial, 2)]

            def cleanUp_consfScores(CAatoms):
                scoresOut = []
                for atom in CAatoms:
                    consfScore = atom.residue.consfScore
                    if consfScore != 999:
                        scoresOut.append(consfScore)
                    elif not consfScore:
                        scoresOut.append('gap')
                    else:
                        scoresOut.append('mismatch')
                return scoresOut
            scoresOut   = cleanUp_consfScores(CAatoms)
            residuesOut = [threeToOneLetter.get(atom.residue.name, atom.residue.name) for atom in CAatoms]

            output = {'indexes':            [atom.residue.number for atom in CAatoms],
                      '%mismatch':          targetChain.percMismatch,
                      '%gap':               targetChain.percGap,
                      'residues':           residuesOut,
                      'consfScores':        scoresOut,
                      'N residue burial':   resBurials[0],
                      'C residue burial':   resBurials[-1],
                      'N residue normalized burial':        resNormBurials[0],
                      'C residue normalized burial':        resNormBurials[-1],}

            structData.update(output)
            return structData
        def find_tail_roots(projDir, targetChain):
            """finds the first buried residue in a root"""
            forwardResidues     = targetChain.residues
            backwardResidues    = forwardResidues[::-1]
            directions          = [forwardResidues, backwardResidues]
            roots = {0:None,1:None}
            for i,direction in enumerate(directions):
                for resi, residue in enumerate(direction):
                    if residue.burial > 10 and 'CA' in gaze_range(AnAtom, residue.atoms, 'name'):
                        roots[i] = residue
                        break

            if not roots[0]:
                issue('failed to find a root for N terminus, this can happen for very small proteins. defaulting to first residue.', projDir)
            if not roots[1]:
                issue('failed to find a root for C terminus, this can happen for very small proteins. defaulting to last residue.', projDir)

            roots[0] = roots[0] if roots[0] else targetChain.residues[0]
            roots[1] = roots[1] if roots[1] else targetChain.residues[-1]
            roots = list(roots.values())
            printWrite(f'roots: {roots}', projDir)

            return roots
        def calculate_interfaceScores(cutoff, atomLocTree, roots, structData):
            def calculate_interfaceScore(targetAtom, cutoff, atomLocTree):
                closeAtoms = find_close_objects(targetAtom, atomLocTree, cutoff)
                closeResidues = set([atom.residue for atom in closeAtoms])
                closeForeignResidues = [residue for residue in closeResidues if
                                        residue.chain != targetAtom.residue.chain]
                distancesToForeignResidues = [get_distance(targetAtom, residue) for residue in closeForeignResidues]
                normalizedResDistances = [cutoff / distance for distance in distancesToForeignResidues]
                interfaceScore = sum([normalizedResDistances if normalizedResDistances else [0]][0])
                return interfaceScore

            rootCAs = fetch(AnAtom, roots, 'name', 'CA')
            NInterfaceScore = calculate_interfaceScore(rootCAs[0], 20, atomLocTree)
            CinterfaceScore = calculate_interfaceScore(rootCAs[1], 20, atomLocTree)
            structData['N interface proximity score'] = NInterfaceScore
            structData['C interface proximity score'] = CinterfaceScore
            return structData
        def find_active_sites(asym_units, distCutoff:int, site_category:int):
            """ Find conserved sites in a structure based on conserved residues and close proximity. generate ActiveSite objects """
            def bin_by_location(residues, cutoff):
                """ Group residues by close locations """
                residueBins = []
                for residue in residues:
                    placed = False
                    for residueBin in residueBins:
                        locationBins = make_LocTree(residueBin)
                        if find_close_objects(residue, locationBins, cutoff):
                            residueBin.append(residue)
                            placed = True
                            break
                    if not placed:
                        residueBins.append([residue])
                return residueBins
            def residue_bin_activeSites(activeSites):
                """
                Group ActiveSites by overlapping residues

                :param activeSites: List of ActiveSites
                """
                activeSiteBins = []
                for activeSite in activeSites:
                    placed = False
                    for bin in activeSiteBins:
                        if any(set(bin[0].residues) & set(activeSite.residues)):
                            bin.append(activeSite)
                            placed = True
                    if not placed:
                        activeSiteBins.append([activeSite])
                return activeSiteBins
            def bins_to_activeSites(bins):
                activeSites = []
                for bin in bins:
                    location        = average_location(bin)
                    sortedResidues  = sorted(bin, key=lambda residue: residue.name)
                    activeSite      = AnActiveSite(sortedResidues, location)
                    activeSites     .append(activeSite)
                return activeSites
            def find_met0(asymUnits):
                """finds name and number of first methionine. Because sometimes Met0 isnt at 1"""
                for i,residue in enumerate(asymUnits[0].chains[0].residues):
                    if residue.name == 'MET':
                        return str(residue)

            conservedResidues   = fetch(AResidue, asym_units, 'colorScore', site_category)  # fetch all conserved residues
            met0                = find_met0(asym_units)
            noMet               = [residue for residue in conservedResidues if str(residue) != met0]
            residueBins         = bin_by_location(noMet, distCutoff)
            cleanedActiveSites  = bins_to_activeSites(residueBins)
            finalActiveSites    = [activeSite for activeSite in cleanedActiveSites if len(activeSite.residues) > 1]

            for activeSite in finalActiveSites:
                for residue in activeSite.residues:
                    residue.sites[site_category] = activeSite

            return finalActiveSites
        def generate_pdb_contents(AnAsymUnits):
            """ Generate a PDB formatted string from transformed asymmetric units. """
            pdb_contents = ""
            atom_serial = 1

            def format_atom_line(atom_serial, atom, residue_name, chain_id, residue_seq):
                """ Format an atom line for a PDB file. """
                # PDB format: Columns 7-11 (atom serial number), 13-16 (atom name),
                # 18-20 (residue name), 22 (chain identifier), 23-26 (residue sequence number),
                # 31-38 (X coordinate), 39-46 (Y coordinate), 47-54 (Z coordinate)
                return f"ATOM  {atom_serial:>5} {atom.name:<4} {residue_name:>3} {chain_id}{residue_seq:>4}    " \
                       f"{atom.location[0]:>8.3f}{atom.location[1]:>8.3f}{atom.location[2]:>8.3f}\n"

            for asym_unit in AnAsymUnits:
                for chain in asym_unit.chains:
                    residue_seq = 1  # Initialize residue sequence number
                    for residue in chain.residues:
                        for atom in residue.atoms:
                            pdb_contents += format_atom_line(atom_serial, atom, residue.name, chain.name, residue_seq)
                            atom_serial += 1
                        residue_seq += 1  # Increment residue sequence number for each residue

            return pdb_contents
        def crawl_cavities(targetChain, AnAsymUnits, structData, atomLocTree, roots, activeSites,
                           conservedSites, pdbID, pdbDir, resetCavities, maxTagDist):
            """finds cavities near termini, crawls them searching for conserved residues"""

            #initialize atoms and residues
            allResidues = fetch(AResidue, AnAsymUnits)
            allAtoms    = fetch(AnAtom, allResidues)
            roots       = fetch(AnAtom, roots, 'name', 'CA')

            #get CavPAtoms and CavPAtomLocTree
            cavPAtomLocFileName = f'{pdbID}_cavPAtomLocs.json'
            if resetCavities:
                delete_from_projDir(cavPAtomLocFileName, projDir)
            def find_cavPAtoms_Locs(allAtoms, atomLocTree, pdbID, pdbDir, roots, maxTagDist):
                """make pseudoAtoms to ocupy cavities, return their LocTree"""


                def find_surfacePAtoms(allAtoms, roots, maxTagDist):
                    """floodfill protein to find pseudoatoms that fill cavities"""
                    #fill full protein area with pseudoAtom locations
                    def initialize_pseudoAtomLocs(allAtoms, roots,maxTagDist):
                        pseudoAtomLocs = set()

                        #generate cube of locations centered on root
                        cubeLocs = []
                        for root in roots:
                            distance        = maxTagDist+15
                            rootlocation    = [int(i) for i in root.location]
                            minMaxes        = [(rootlocation[dimension]-distance, rootlocation[dimension]+distance) for
                                               dimension in range(3)]
                            dimensionLocs   = [set([x for x in range(minMaxes[i][0], minMaxes[i][1])]) for i in range(3)]
                            for x in dimensionLocs[0]:
                                for y in dimensionLocs[1]:
                                    for z in dimensionLocs[2]:
                                        pseudoAtomLoc = (x,y,z)
                                        cubeLocs.append(pseudoAtomLoc)

                            #exclude points outside of a sphere of same radius
                            cubeLocsTree    = make_LocTree(cubeLocs)
                            sphereLocs      = find_close_objects(root, cubeLocsTree, distance)
                            pseudoAtomLocs  .update(sphereLocs)

                        return list(pseudoAtomLocs)
                    pAtomLocs      = initialize_pseudoAtomLocs(allAtoms, roots,maxTagDist)
                    pAtomLocTree     = make_LocTree(pAtomLocs)

                    #exclude pAtomLocs that are too close to real atoms
                    buriedPAtomLocBins  = [find_close_objects(atom, pAtomLocTree, 2.3) for atom in allAtoms]
                    buriedPAtomLocs     = set(atom for bin in buriedPAtomLocBins for atom in bin)
                    freePAtomLocs       = set(pAtomLocs).difference(buriedPAtomLocs)

                    #exclude pAtomLocs that are too far from real atoms
                    surfacePAtomLocs      = set(pAtomLoc for pAtomLoc in freePAtomLocs if
                                                len(find_close_objects(pAtomLoc, atomLocTree, 5, isArray = True))>5)

                    #convert locations to pAtoms
                    surfacePAtoms = [AnAtom(location=np.array(pAtomLoc)) for pAtomLoc in surfacePAtomLocs]
                    return surfacePAtoms
                surfacePAtoms = find_surfacePAtoms(allAtoms, roots,maxTagDist)

                #assign burials to pAtoms
                def assign_pAtom_normBurials(allAtoms, atomLocTree, burialRadius):
                    """calculating isCavity requires calculating burials for all atoms a second time, at 14A
                    the difference between the normalized atom number at 5 vs 14 angstroms is used to determine cavity status"""

                    #calc burials
                    atomBurials       = _calc_burials(allAtoms, atomLocTree, burialRadius)

                    #normalize burials
                    normAtomBurials = normalize_aList(atomBurials)

                    #assign normalized burials
                    for atomi, atom in enumerate(allAtoms):
                        atom.normBurials[burialRadius] = normAtomBurials[atomi]
                assign_pAtom_normBurials(surfacePAtoms, atomLocTree, 5)
                assign_pAtom_normBurials(surfacePAtoms, atomLocTree, 14)

                #determine cavity status of pAtoms using their burials
                def get_cavity_pAtoms(allPAtoms, burLowerBound, burUpperBound, burDifference, minBurLower):
                    """returns a list of pAtoms that are in cavities, based on normalized burials at two distances"""

                    #cavityAtomNumbers = set()
                    cavPAtoms = set()
                    for atom in allPAtoms:
                        isCavity      = (atom.normBurials[burUpperBound] - atom.normBurials[burLowerBound] > 0.4 and
                                         atom.normBurials[burLowerBound] < 0.4)
                        if isCavity:
                            cavPAtoms.add(atom)
                            #cavityAtomNumbers.add(atom.number)
                    return list(cavPAtoms)
                cavPAtoms        = get_cavity_pAtoms(surfacePAtoms, 5, 14, 0.4, 0.4)

                #save cavPAtoms to avoid future floodfills
                cavPAtomLocs = [[int(dimension) for dimension in cavPAtom.location] for cavPAtom in cavPAtoms]
                return cavPAtomLocs

            allAtoms        = fetch(AnAtom, AnAsymUnits)
            cavPAtomLocs    = if_not_in_projDir(find_cavPAtoms_Locs, cavPAtomLocFileName, targetDir= pdbDir,
                                                exactName=True, allAtoms=allAtoms, atomLocTree=atomLocTree,
                                                pdbID=pdbID, roots=roots, maxTagDist=maxTagDist, pdbDir = pdbDir)
            cavPAtoms       = [AnAtom(location=location, name = f'pAtom {i}') for i,location in enumerate(cavPAtomLocs)]
            cavPAtomLocTree = make_LocTree(cavPAtoms)

            #make ACavities
            def make_ACavities(cavPAtoms, cavPAtomLocTree, atomLocTree, cutoff):
                """collects pAtoms into cavity objects based on proximity"""

                def bin_pAtoms_by_location(cavPAtoms, cavPAtomLocTree, cutoff):
                    """collects pAtoms within cutoff of each other into individual cavities"""

                    assignedPAtoms  = set()
                    pAtomCavityBins = []
                    for pAtom in cavPAtoms:
                        if pAtom not in assignedPAtoms:
                            binned = False
                            siblingPAtoms = find_close_objects(pAtom, cavPAtomLocTree, cutoff)
                            if siblingPAtoms:
                                for pAtom in siblingPAtoms:
                                    for bin in pAtomCavityBins:
                                        if pAtom in bin:
                                            bin.update(siblingPAtoms)
                                            binned = True
                                            break
                                if not binned:
                                    pAtomCavityBins.append(siblingPAtoms)
                                assignedPAtoms.update(siblingPAtoms)

                    def join_shared_bins(pAtomCavityBins):
                        """Joins bins that share pAtoms"""
                        joined_bins = []
                        for bin1 in pAtomCavityBins:
                            new_bin = bin1.copy()
                            for bin2 in joined_bins:
                                if bin1.intersection(bin2):
                                    new_bin.update(bin2)
                                    joined_bins.remove(bin2)
                            joined_bins.append(new_bin)
                        return joined_bins
                    pAtomCavityBins = join_shared_bins(pAtomCavityBins)

                    return pAtomCavityBins
                pAtomCavityBins = bin_pAtoms_by_location(cavPAtoms, cavPAtomLocTree, cutoff)

                ACavities = []
                for bin in pAtomCavityBins:
                    newACavity = ACavity(bin, atomLocTree)
                    ACavities.append(newACavity)
                return ACavities
            ACavities = make_ACavities(cavPAtoms, cavPAtomLocTree, atomLocTree, 2.5)

            #crawl cavities near each terminus
            rootCavitiesScores  = []

            for root in roots:
                #find cavities near root
                nearRootAtoms               = find_close_objects(root, atomLocTree, maxTagDist)
                nearRootSurfaceAtoms        = set(fetch(AnAtom, nearRootAtoms, "normBurial", .35, "<"))
                nearRootCavities            = list(set(surfaceAtom.cavity for surfaceAtom in nearRootSurfaceAtoms if surfaceAtom.cavity))
                nearCavitySurfaceAtoms      = [nearRootSurfaceAtoms.intersection(cavity.atoms) for cavity in nearRootCavities]
                test_nearCavSurfAtoms       = '+'.join([str(atom.number) for bin in nearCavitySurfaceAtoms for atom in bin])

                def print_cavAtomLine(nearCavitySurfaceAtoms):
                    testAllNearCavSurfAtoms     = [atom for bin in nearCavitySurfaceAtoms for atom in bin]
                    cavAtomLine                 = '+'.join([str(atom.number) for atom in testAllNearCavSurfAtoms])
                    print(f'sele cavAtoms, ID {cavAtomLine}')

                minDistancesToCavities      = [min([get_distance(root, atom) for atom in surfaceAtoms]) for
                                               surfaceAtoms in nearCavitySurfaceAtoms]
                cavProbabs                  = [1-(distance/maxTagDist) for distance in minDistancesToCavities]
                rootCavities                = {cavProbabs[i]:cavity for i, cavity in enumerate(nearRootCavities)}

                rootCavityScores            = []
                for probability, cavity in rootCavities.items():
                    if cavity.volume > 90:
                        rootAccessibleVolume = len(find_close_objects(root, cavity.pAtomLocTree, maxTagDist+5))
                        if rootAccessibleVolume > 90:
                            cavityParameters = {
                                'consfScore': cavity.consfScore,
                                'activeSiteScore': cavity.activeSiteScore,
                                'conservedSiteScore': cavity.conservedSiteScore,
                                'volume': rootAccessibleVolume,
                                'probability': probability,
                            }
                            test_pAtom_locations = [pAtom.location for pAtom in cavity.pAtoms]

                            if cavityParameters:
                                rootCavityScores.append(cavityParameters)
                            else:
                                rootCavityScores.append('no accessible cavities')
                rootCavitiesScores.append(rootCavityScores)

            results = { 'N cavities': rootCavitiesScores[0],
                        'C cavities': rootCavitiesScores[1]}

            structData.update(results)
            return structData
        def print_consurfColorCommand(AnAsymUnits):
            """prints a command to separate residues into colored pyMol selections by conservation"""
            residues        = fetch(AResidue, AnAsymUnits)

            #sort resNums by conservation
            conservations   = {i:set() for i in range(1,10)}
            conservations[999] = set()

            for residue in residues:
                colorScore                = residue.colorScore
                resNum                    = residue.number
                if resNum != None and resNum > 0:  #color pyMol seems to reinterpret negatives as positive
                    conservations[colorScore] .add(resNum)
            #generate color commands
            colors      = ['0xf9ffcf', '0xf3bfc0', '0xed80b1', '0xe840a3', '0xe20094', '0xb20084', '0x820074', '0x530063',
                            '0x230053', '0x32CD32']

            sequences   = ['+'.join([str(resNum) if resNum>0 else None for resNum in conservations[colorScore]]) for colorScore in conservations]
            for i, sequence in enumerate(sequences):
                print('sele ' + str(i + 1) + ', resi ' + sequences[i])
                print('color', colors[i] + ', resi ' + sequences[i])

        warnings.filterwarnings("ignore")

        pdbFileName            = name_from_projDir(f'{pdbID}.pdb', pdbDir, exactName=True)                                   #get pdb file name
        origAsyms, structData  = make_asym_units_get_index_map(
                                            pdbFileName, projDir, pdbDir, targetSeq, pdbID, consfDicts, structData)              #make asymUnits, retrieve index map
        if not origAsyms or not any(gaze_range(AChain, origAsyms, 'goodChain')):
            issue(f'No valid chains found in {pdbID}', projDir)
            output = {'indexes'     : [1],
                      'residues'    : ['failed to find valid chains'],
                      'consfScores' : ['failed to find valid chains']}
            structData.update(output)
            return structData
        transformMatrices, oligo, structData = extract_matrices_oligo_gmqe_qsqe(projDir, pdbFileName, structData)                #extract matrices and oligomer
        AnAsymUnits                          = transform_asyms(origAsyms, transformMatrices, oligo)                     #generate or rename quaternary structure
        targetChain, structData              = find_targetChain(AnAsymUnits, targetChainName, structData)               #find target chain
        atomLocTree                          = make_LocTree(fetch(AnAtom, AnAsymUnits))                                 #make position tree for all atoms
        residueLocTree                       = make_LocTree(fetch(AResidue, AnAsymUnits))                               #make position tree for all residues
        structData                           = extract_unresolved_termini(targetChain, structData)
        structData                           = calculate_burials(targetChain,6,AnAsymUnits, structData, atomLocTree)             #calculate burial scores
        roots                                = find_tail_roots(projDir, targetChain)
        structData                           = calculate_interfaceScores(20, atomLocTree, roots, structData)
        activeSites                          = find_active_sites(AnAsymUnits, distCutoff=6, site_category=9)               #find active sites
        conservedSites                       = find_active_sites(AnAsymUnits, distCutoff=6, site_category=8)               #find conserved sites
        if cavitySearch:
            printWrite('searching cavities', projDir)
            structData                           = crawl_cavities(targetChain, AnAsymUnits, structData,
                                                                  atomLocTree, roots, activeSites,
                                                                  conservedSites, pdbID, pdbDir, resetCavities, maxTagDist=20)

        if getColorCmds:
            print_consurfColorCommand(AnAsymUnits)

        return structData
    def consurf_to_output(consfDicts, structData):
        """convert consurf output to a dictionary for output with results"""
        consfScores, _, consfPositions, errorDict = consfDicts if consfDicts else\
            [{None:None} for _ in range(3)]+[{'%mismatch':None,'%gap':None}]

        structData['%mismatch'] = errorDict['%mismatch']
        structData['%gap']      = errorDict['%gap']

        indexes                 = list(consfScores.keys())
        indexes                 . remove(None)
        structData['indexes']   = indexes

        residues                = [consfPositions[index] for index in indexes]
        structData['residues']  = residues

        scores                      = [consfScores[index] for index in indexes]
        structData['consfScores']   = scores

        return structData
    def uniprot_to_output(targetSeq, uniprotID, structData):
        """convert consurf output to a dictionary for output with results"""

        indexes      = [i for i in range(1, len(targetSeq)+1)]
        residues     = targetSeq

        structData['indexes']     = indexes
        structData['residues']    = residues
        structData['consfScores'] = residues

        return structData

    structure_datas   = {}
    for i, pdbID in enumerate(all_pdbIDs):

        startTime = time.time()
        printWrite(f'\n>{pdbID} structure {i+1}/{len(all_pdbIDs)}', projDir)
        structData = deepcopy(defaultStructData)
        try:
            structure_datas[pdbID]  = process_structure(pdbID, projDir, pdbDir, consfDicts, structData,
                                                        defaultStructData, targetChainName, resetCavities)
            printWrite(f'{pdbID} completed in {round(time.time()-startTime,2)} seconds', projDir)
        except:
            traceback = get_traceback()
            issue(f'Error processing {pdbID} with error:\n{traceback}', projDir)
    structure_datas[f'CONSURF_{consfID}'] = consurf_to_output(consfDicts, deepcopy(defaultStructData))
    structure_datas[f'uniprot_{uniprotID}'] = uniprot_to_output(targetSeq, uniprotID, deepcopy(defaultStructData))
    return structure_datas
def process_structure_datas(structure_datas, defaultStructData, projDir, uniprotTitle, sheetName):
    """ Process structure data and save to Excel. """
    def export_scores(projDir, outputFilePath, df, outputFileName, sheetName):
        def color_cells_by_amino_acids(excelFilePath, sheetName):
            amino_acid_colors = {
                'A': '7FFFD4',  # Alanine - Aquamarine
                'R': '0000FF',  # Arginine - Blue
                'N': '8A2BE2',  # Asparagine - Blue Violet
                'D': 'A52A2A',  # Aspartic Acid - Brown
                'C': 'DEB887',  # Cysteine - Burlywood
                'E': '5F9EA0',  # Glutamic Acid - Cadet Blue
                'Q': 'D2691E',  # Glutamine - Chocolate
                'G': '00FFFF',  # Glycine - Cyan
                'H': 'FF7F50',  # Histidine - Coral
                'I': 'DC143C',  # Isoleucine - Crimson
                'L': '00FF00',  # Leucine - Lime
                'K': '8B008B',  # Lysine - Dark Magenta
                'M': '556B2F',  # Methionine - Dark Olive Green
                'F': 'FF1493',  # Phenylalanine - Deep Pink
                'P': '00BFFF',  # Proline - Deep Sky Blue
                'S': '696969',  # Serine - Dim Gray
                'T': '1E90FF',  # Threonine - Dodger Blue
                'W': 'B8860B',  # Tryptophan - Dark Goldenrod
                'Y': '2E8B57',  # Tyrosine - Sea Green
                'V': 'DAA520',  # Valine - Goldenrod
            }
            workbook    = load_workbook(excelFilePath)
            sheet       = workbook[sheetName]

            for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
                for cell in row:
                    if cell.value in amino_acid_colors:
                        # Apply fill color
                        fill = PatternFill(start_color=amino_acid_colors[cell.value],
                                           end_color=amino_acid_colors[cell.value], fill_type="solid")
                        cell.fill = fill
                    cell.alignment = Alignment(horizontal='left')

            # Save the changes
            workbook.save(excelFilePath)
        if sheetName == 'residues':
            with pd.ExcelWriter(outputFilePath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheetName)
            color_cells_by_amino_acids(outputFilePath, sheetName)
        else:
            with pd.ExcelWriter(outputFilePath, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheetName)
        printWrite(f'{outputFileName} sheet {sheetName} saved to {projDir}', projDir)

    if not structure_datas:
        noDataMsg = """OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
        None of the extracted crystal structures produced processable data.
        This is usually because the consurf result fetched was not appropriate for this protein,
        due to crystal structures focusing on protein-protein interfaces containing chains from different proteins.
        
        delete consurf_summary.txt in the project directory,
        Visit the RCSB pages for the PDB IDs listed above,
        under "macromolecules" find the chain which contains your protein,
        and set consurfPDB and consurfChain accordingly.
        OOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO"""
        issue(noDataMsg, projDir)

    outputFileName = f'{uniprotTitle}_termini_scores.xlsx'
    outputFilePath = name_from_projDir(outputFileName, projDir)


    # Collect all unique x-values (indexes) across all IDs and sort them
    all_x_values = sorted(set(x for data in structure_datas.values() for x in data['indexes'] if x != None))

    # Initialize aligned_scores with additional keys for metadata
    sheetDatas = {'pdbID':[]}
    sheetDatas .update({key:[] for key, _ in defaultStructData.items() if key not in ['indexes', 'residues', 'consfScores']})
    sheetDatas .update({' ': []})
    sheetDatas .update({x: [] for x in all_x_values})


    # Process each ID and their associated data
    for ID, data in structure_datas.items():
        # Extract data for the current ID
        if data:
            indexes                     = data['indexes']
            if sheetName == 'residues':
                residues                    = data['residues']
            elif sheetName == 'consfScores':
                residues                    = data['consfScores']


            # Create a row for each index and score, filling missing scores with None
            sheetDatas['pdbID']         .append(ID)
            sheetDatas[' ']             .append('')
            for sheetIndex in all_x_values:
                sheetDatas[sheetIndex].append(residues[indexes.index(sheetIndex)] if sheetIndex in indexes else None)
            # Append the values to their corresponding keys in aligned_scores
            for key, value in data.items():
                if key not in ['indexes', 'residues', 'consfScores']:
                    if value != 999:
                        sheetDatas[key].append(value)
                    else:
                        sheetDatas[key].append('')

    # Convert to DataFrame, transpose, and set 'ID' as the index
    df = pd.DataFrame.from_dict(sheetDatas).set_index('pdbID')
    df = df.transpose()

    # Save to Excel with error handling
    try:
        export_scores(projDir, outputFilePath, df, outputFileName, sheetName)
    except PermissionError:
        input(f'Failed to save {outputFileName}. If it is open in another program, close it and press ENTER.')
        export_scores(projDir, outputFilePath, df, outputFileName, sheetName)
    except Exception:
        printWrite(f'Failed to save {outputFileName}. error: {traceback.format_exc()}', projDir)



